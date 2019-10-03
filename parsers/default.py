import asyncio
import datetime

from openpyxl import load_workbook

from db import Group, Schedule, prepare_db


class TableParser:
    def __init__(self, path_to_file):
        self.docs = load_workbook(path_to_file)
        self.table = self.docs.active
        self._merged_cells_fill()
        self.groups = []

    def _merged_cells_fill(self):
        for crange in self.table.merged_cells:
            for row in range(crange.min_row, crange.max_row + 1):
                for col in range(crange.min_col, crange.max_col + 1):
                    self.table.cell(row, col).value = self.table.cell(crange.min_row, crange.min_col).value

    async def parse(self):
        for table_name in self.docs.sheetnames:
            print(f"parse: {table_name}")
            self.table = self.docs.get_sheet_by_name(table_name)
            self._merged_cells_fill()
            row = 13

            while True:
                col = 2

                for i in range(7):  # Try find next week
                    date = self.table.cell(row, col).value
                    group = self.table.cell(row + 2, col).value

                    # print(date, group)
                    if not (date and group):
                        row += 1
                    else:
                        # print("Skipped {} row. Find {}-{}".format(i, date, group))
                        break
                else:
                    break

                while True:
                    date = self.table.cell(row, col).value
                    group = self.table.cell(row + 2, col).value

                    if not date or not group:
                        break
                    group_id = (await Group.get_or_create(group, table_name)).id
                    for i in range(1, 9):
                        information = self.table.cell(row + 2 + i, col).value
                        if information:
                            await Schedule.create(group_id, date, i, information)

                    col += 1
                row += 12
            print(f"parsed: {table_name}")


async def run():
    await prepare_db()

    parser = TableParser('../data/table.xlsx')
    await parser.parse()


if __name__ == '__main__':
    asyncio.run(run())
