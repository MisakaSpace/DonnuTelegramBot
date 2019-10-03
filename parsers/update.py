import asyncio
import logging

LOGGER = logging.getLogger(__name__)

from openpyxl import load_workbook

from db import Group, Schedule, prepare_db


class TableParser:
    def __init__(self, path_to_file):
        self.docs = load_workbook(path_to_file)
        self.table = self.docs.active
        self._merged_cells_fill()
        self.groups = []

    def _merged_cells_fill(self):
        for crange in self.table.merged_cells:
            for row in range(crange.min_row, crange.max_row + 1):
                for col in range(crange.min_col, crange.max_col + 1):
                    self.table.cell(row, col).value = self.table.cell(crange.min_row, crange.min_col).value

    async def parse(self, check=False):
        if not check:
            await Schedule.clear()
        for table_name in self.docs.sheetnames:
            LOGGER.info(f"Start parse: {table_name}")
            self.table = self.docs.get_sheet_by_name(table_name)

            self._merged_cells_fill()
            row = 13

            while True:
                col = 2

                date = self.table.cell(row, col).value
                group = self.table.cell(row + 2, col).value
                if not (date and group):
                    row += 1
                    date = self.table.cell(row, col).value
                    if not (date and group) or date == 'пн':
                        break

                while True:
                    date = self.table.cell(row, col).value
                    group = self.table.cell(row + 2, col).value
                    if not date or not group:
                        break
                    group_id = (await Group.get_or_create(group, table_name)).id
                    for i in range(1, 9):
                        information = self.table.cell(row + 2 + i, col).value
                        if information:
                            if not check:
                                await Schedule.create(group_id, date, i, information)

                    col += 1
                row += 12
            LOGGER.info(f"End parse: {table_name}")


async def run():
    await prepare_db()

    parser = TableParser('../data/table.xlsx')
    await parser.parse()


if __name__ == '__main__':
    asyncio.run(run())
